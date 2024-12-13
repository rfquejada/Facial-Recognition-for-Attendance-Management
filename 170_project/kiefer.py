import face_recognition
import cv2
import os
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Load known faces
known_faces_dir = "known_faces"
known_encodings = []
known_names = []

for person_name in os.listdir(known_faces_dir):
    person_dir = os.path.join(known_faces_dir, person_name)
    if not os.path.isdir(person_dir):
        continue

    print(f"Processing {person_name}'s images...")
    for image_name in os.listdir(person_dir):
        image_path = os.path.join(person_dir, image_name)
        # Load the image and encode the face
        image = face_recognition.load_image_file(image_path)
        encodings = face_recognition.face_encodings(image)
        if encodings:
            # Add the encoding and the person's name to the lists
            known_encodings.append(encodings[0])
            known_names.append(person_name)

print(f"Loaded {len(known_encodings)} faces from {len(set(known_names))} people.")

# Create or load the attendance file
attendance_file = "attendance.xlsx"
if not os.path.exists(attendance_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name"] + [datetime.now().strftime("%Y-%m-%d")])
    for name in set(known_names):
        ws.append([name])
    wb.save(attendance_file)

# Initialize the camera
camera = cv2.VideoCapture(0)

if not camera.isOpened():
    print("Error: Could not open the camera.")
    exit()

print("Press 'q' to quit.")

# List to track who has been marked present
marked_present = set()

while True:
    # Capture a frame
    ret, frame = camera.read()
    if not ret:
        print("Failed to grab a frame.")
        break

    # Resize frame for faster processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    # Detect faces in the frame
    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        # Scale back up face locations to match the original frame size
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        # Compare the detected face to the known faces
        matches = face_recognition.compare_faces(known_encodings, face_encoding)
        name = "Unknown"

        if True in matches:
            # Find the first match index
            match_index = matches.index(True)
            name = known_names[match_index]

            if name not in marked_present:
                wb = load_workbook(attendance_file)
                ws = wb["Attendance"]
                today = datetime.now().strftime("%Y-%m-%d")
                
                # Extract headers and ensure "TIME" column is present
                headers = [cell.value for cell in ws[1] if cell.value is not None]
                
                # Add the date column if missing
                if today not in headers:
                    ws.cell(row=1, column=len(headers) + 1, value=today)
                    headers.append(today)
                
                # Add the "TIME" column if missing
                if "Time" not in headers:
                    ws.cell(row=1, column=len(headers) + 1, value="Time")
                    headers.append("Time")

                # Mark as present and record time
                current_time = datetime.now().strftime("%H:%M:%S")
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                    if row[0].value == name:
                        date_col_index = headers.index(today) + 1
                        time_col_index = headers.index("Time") + 1
                        ws.cell(row=row[0].row, column=date_col_index, value="PRESENT")
                        ws.cell(row=row[0].row, column=time_col_index, value=current_time)
                        break

                wb.save(attendance_file)
                marked_present.add(name)

        # Draw a rectangle around the face
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

        # Display the name below the rectangle
        cv2.putText(frame, name, (left, bottom + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 255, 0), 2)

    # Display the video feed
    cv2.imshow("Camera", frame)

    # Exit the loop if 'q' is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the camera and close all windows
camera.release()
cv2.destroyAllWindows()
