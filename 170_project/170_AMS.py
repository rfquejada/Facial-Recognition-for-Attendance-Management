import face_recognition
import cv2
import os
import numpy as np
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Load known faces
known_faces_dir = "known_faces"
known_encodings = []
known_names = []

# Track True Positives, False Positives, and False Negatives
true_positives = 0
false_positives = 0
false_negatives = 0
total_faces = 0  # Total number of faces processed

for person_name in os.listdir(known_faces_dir):
    person_dir = os.path.join(known_faces_dir, person_name)
    if not os.path.isdir(person_dir):
        continue

    print(f"Processing {person_name}'s images...")  
    for image_name in os.listdir(person_dir):   #access the directory containing the person name
        image_path = os.path.join(person_dir, image_name)
        image = face_recognition.load_image_file(image_path)  #load the image of the person for processing
        encodings = face_recognition.face_encodings(image)
        if encodings:
            #add the encoding and the person's name to the lists
            known_encodings.append(encodings[0])
            known_names.append(person_name)

print(f"Loaded {len(known_encodings)} faces from {len(set(known_names))} people.")

# Sort the names alphabetically
known_names_sorted = sorted(set(known_names))
                            
#create or load the attendance file for appending or creating
attendance_file = "attendance.xlsx"
if not os.path.exists(attendance_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name"] + [datetime.now().strftime("%Y-%m-%d")])
    for name in known_names_sorted:
        # Initialize each name with ABSENT for the current date
        ws.append([name, "ABSENT"])
    wb.save(attendance_file)

# Mark attendance with default 'ABSENT' for new dates
else:
    wb = load_workbook(attendance_file)
    ws = wb["Attendance"]
    today = datetime.now().strftime("%Y-%m-%d")
    headers = [cell.value for cell in ws[1] if cell.value is not None]

    # Add the date column if missing and set default to 'ABSENT'
    if today not in headers:
        ws.cell(row=1, column=len(headers) + 1, value=today)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers) + 1):
            row[-1].value = "ABSENT"
    wb.save(attendance_file)

# initialize the camera
camera = cv2.VideoCapture(0)

if not camera.isOpened():  #check if the camera is open or not
    print("Error: Could not open the camera.")
    exit()

print("Press 'q' to quit.")   #check if the attendance process is finished or done

# List to track who has been marked present
marked_present = set()

while True:
    # Capture a frame
    ret, frame = camera.read()
    if not ret:
        print("Failed to grab a frame.")
        break

    #resize the frame for accurate processing of result of the person
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    #detect the faces of the person in the given frame considering the database of the students
    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        total_faces += 1  #increment the total faces processed

        #scale back up face locations to match the original frame size
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        #comparison of the detected faces on the student database
        distances = face_recognition.face_distance(known_encodings, face_encoding)
        best_match_index = np.argmin(distances)
        name = "Unknown"

        if distances[best_match_index] < 0.6:  #threshold for a confident match
            name = known_names[best_match_index]

            if name not in marked_present:
                wb = load_workbook(attendance_file)
                ws = wb["Attendance"]
                today = datetime.now().strftime("%Y-%m-%d")  #get the current time for the attendance
                
                # Extract headers and ensure "Time" column is present
                headers = [cell.value for cell in ws[1] if cell.value is not None]
                
                #add the date
                if today not in headers:
                    ws.cell(row=1, column=len(headers) + 1, value=today)
                    headers.append(today)
                
                #add the time of attendance
                if "Time" not in headers:
                    ws.cell(row=1, column=len(headers) + 1, value="Time")
                    headers.append("Time")

                #mark the student as present and record the time 
                current_time = datetime.now().strftime("%H:%M:%S")
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                    if row[0].value == name:
                        date_col_index = headers.index(today) + 1
                        time_col_index = headers.index("Time") + 1
                        ws.cell(row=row[0].row, column=date_col_index, value="PRESENT")
                        ws.cell(row=row[0].row, column=time_col_index, value=current_time)
                        break

                wb.save(attendance_file)
                marked_present.add(name)

                true_positives += 1  #correctly identify the person, mark as true positive

            else:
                false_positives += 1  # False alarm for someone already marked as present

        else:
            false_negatives += 1  # Face detected but did not match known faces

        #create a frame for visualization
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

        #display the name for the student
        cv2.putText(frame, name, (left, bottom + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 255, 0), 2)

    #display the video feed
    cv2.imshow("Camera", frame)

    # Exit the loop if 'q' is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Calculate performance metrics after the loop ends
accuracy = (true_positives / total_faces) * 100 if total_faces > 0 else 0
precision = (true_positives / (true_positives + false_positives)) * 100 if (true_positives + false_positives) > 0 else 0
recall = (true_positives / (true_positives + false_negatives)) * 100 if (true_positives + false_negatives) > 0 else 0

# Output the results
print("\nPerformance Metrics:")
print(f"Total Faces Processed: {total_faces}")
# print(f"True Positives: {true_positives}")
# print(f"False Positives: {false_positives}")
# print(f"False Negatives: {false_negatives}")
# print(f"Accuracy: {accuracy:.2f}%")
# print(f"Precision: {precision:.2f}%")
# print(f"Recall: {recall:.2f}%")

# Release the camera and close all windows
camera.release()
cv2.destroyAllWindows()
